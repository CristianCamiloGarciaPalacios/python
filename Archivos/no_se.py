import openpyxl

excel_document = openpyxl.load_workbook('Archivos/ejemplo.xlsx')

print(type(excel_document)) #Tipo de archivo

print(excel_document.sheetnames) #Nombre de las hojas del archivo exel

sheet = excel_document['Hoja1'] #Agregar a una variable el contenido de una hoja

print(sheet['A2'].value) #Imprimir el valor de una celda de la hoja, a partir del nombre de la celda

print(sheet.cell(row = 5, column = 2).value) #Imprimir el valor de una celda de la hoja, a partir de sus coordenadas

print(type(sheet['A2'])) #Tipo de archivo que es la celda

print(sheet.cell(row = 5, column = 2)) #Imprimir propoiedades de una celda

multiple_cells = sheet['A1':'G4'] #Valor de un rango de celdas (forma de matiz: a1 = [0][0]; b3 = [1][2])
for row in multiple_cells:
    for cell in row:
        print(cell.value," ", end="")
    print("")

all_rows = sheet.rows
for row in all_rows:
    for cell in row:
        print(cell.value," ", end="")
    print("")